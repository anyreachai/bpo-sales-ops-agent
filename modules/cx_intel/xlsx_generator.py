"""Generate a professional CX Intelligence XLSX workbook from scraped review data.

5-sheet workbook: Executive Summary, Theme Analysis, Consumer Reviews,
Employee Reviews, Recommendations.
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Design System ─────────────────────────────────────────────────────

ZINC_900 = "18181B"
ZINC_800 = "27272A"
ZINC_700 = "3F3F46"
ZINC_600 = "52525B"
ZINC_400 = "A1A1AA"
ZINC_200 = "E4E4E7"
ZINC_100 = "F4F4F5"
ZINC_50 = "FAFAFA"
WHITE = "FFFFFF"

GREEN_700 = "15803D"
GREEN_100 = "DCFCE7"
AMBER_700 = "B45309"
AMBER_100 = "FEF3C7"
RED_700 = "B91C1C"
RED_100 = "FEE2E2"
BLUE_700 = "1D4ED8"
BLUE_100 = "DBEAFE"

DEFAULT_ACCENT = "5B5FC7"

# Fills
FILL_ZINC_900 = PatternFill(start_color=ZINC_900, end_color=ZINC_900, fill_type="solid")
FILL_ZINC_800 = PatternFill(start_color=ZINC_800, end_color=ZINC_800, fill_type="solid")
FILL_ZINC_100 = PatternFill(start_color=ZINC_100, end_color=ZINC_100, fill_type="solid")
FILL_ZINC_50 = PatternFill(start_color=ZINC_50, end_color=ZINC_50, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

FILL_GREEN = PatternFill(start_color=GREEN_100, end_color=GREEN_100, fill_type="solid")
FILL_AMBER = PatternFill(start_color=AMBER_100, end_color=AMBER_100, fill_type="solid")
FILL_RED = PatternFill(start_color=RED_100, end_color=RED_100, fill_type="solid")
FILL_BLUE = PatternFill(start_color=BLUE_100, end_color=BLUE_100, fill_type="solid")

# Fonts
FONT_TITLE = Font(name="Calibri", bold=True, color=WHITE, size=16)
FONT_SECTION = Font(name="Calibri", bold=True, color=WHITE, size=12)
FONT_HEADER = Font(name="Calibri", bold=True, color=ZINC_800, size=10)
FONT_BODY = Font(name="Calibri", color=ZINC_700, size=10)
FONT_BODY_BOLD = Font(name="Calibri", bold=True, color=ZINC_800, size=10)
FONT_SMALL = Font(name="Calibri", color=ZINC_600, size=9)
FONT_KPI_VALUE = Font(name="Calibri", bold=True, color=ZINC_900, size=22)
FONT_KPI_LABEL = Font(name="Calibri", color=ZINC_600, size=9)
FONT_WHITE_SM = Font(name="Calibri", bold=True, color=WHITE, size=10)

FONT_GREEN = Font(name="Calibri", bold=True, color=GREEN_700, size=10)
FONT_AMBER = Font(name="Calibri", bold=True, color=AMBER_700, size=10)
FONT_RED = Font(name="Calibri", bold=True, color=RED_700, size=10)
FONT_BLUE = Font(name="Calibri", bold=True, color=BLUE_700, size=10)

# Borders
BORDER_BOTTOM = Border(bottom=Side(style="thin", color=ZINC_200))
BORDER_ALL = Border(
    left=Side(style="thin", color=ZINC_200),
    right=Side(style="thin", color=ZINC_200),
    top=Side(style="thin", color=ZINC_200),
    bottom=Side(style="thin", color=ZINC_200),
)

# Alignments
ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _sentiment_style(sentiment: str) -> tuple[PatternFill, Font]:
    s = (sentiment or "").lower()
    if s == "positive":
        return FILL_GREEN, FONT_GREEN
    if s == "negative":
        return FILL_RED, FONT_RED
    return FILL_AMBER, FONT_AMBER


def _frequency_style(freq: str) -> tuple[PatternFill, Font]:
    f = (freq or "").lower()
    if f == "high":
        return FILL_RED, FONT_RED
    if f == "medium":
        return FILL_AMBER, FONT_AMBER
    return FILL_GREEN, FONT_GREEN


def _set_col_widths(ws, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _accent_fill(accent: str) -> PatternFill:
    return PatternFill(start_color=accent.lstrip("#"), end_color=accent.lstrip("#"), fill_type="solid")


def _write_title_bar(ws, row: int, text: str, col_span: int, fill: PatternFill = FILL_ZINC_900) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_TITLE
    cell.fill = fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 36
    for c in range(2, col_span + 1):
        ws.cell(row=row, column=c).fill = fill
    return row + 1


def _write_section_header(ws, row: int, text: str, col_span: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_SECTION
    cell.fill = FILL_ZINC_800
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28
    for c in range(2, col_span + 1):
        ws.cell(row=row, column=c).fill = FILL_ZINC_800
    return row + 1


def _write_header_row(ws, row: int, headers: list[str]) -> int:
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_ZINC_100
        cell.border = BORDER_ALL
        cell.alignment = ALIGN_CENTER
    ws.row_dimensions[row].height = 24
    return row + 1


def _write_cell(ws, row: int, col: int, value, font=FONT_BODY, alignment=ALIGN_LEFT_TOP, fill=None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.border = BORDER_BOTTOM
    cell.alignment = alignment
    if fill:
        cell.fill = fill


def _alt_row_fill(row_idx: int) -> PatternFill | None:
    return FILL_ZINC_50 if row_idx % 2 == 0 else None


# ======================================================================
# Sheet builders
# ======================================================================


def _build_executive_summary(wb: Workbook, data: dict, company_name: str, accent: str) -> None:
    ws = wb.active
    ws.title = "Executive Summary"
    _set_col_widths(ws, {1: 24, 2: 18, 3: 18, 4: 18, 5: 18, 6: 30})

    afill = _accent_fill(accent)

    row = _write_title_bar(ws, 1, f"CX Intelligence Report — {company_name}", 6, afill)
    row += 1  # spacer

    # KPI cards row
    overall = data.get("overall_rating")
    total = data.get("total_reviews_found", 0)
    dist = data.get("sentiment_distribution", {})
    total_sentiment = sum(dist.get(k, 0) for k in ("positive", "mixed", "negative")) or 1
    positive_pct = round(100 * dist.get("positive", 0) / total_sentiment)
    negative_pct = round(100 * dist.get("negative", 0) / total_sentiment)

    kpis = [
        (str(overall) if overall is not None else "N/A", "Overall Rating"),
        (str(total), "Total Reviews"),
        (f"{positive_pct}%", "Positive Sentiment"),
        (f"{negative_pct}%", "Negative Sentiment"),
    ]

    for i, (val, label) in enumerate(kpis):
        col = i + 1
        cell_v = ws.cell(row=row, column=col, value=val)
        cell_v.font = FONT_KPI_VALUE
        cell_v.alignment = ALIGN_CENTER
        cell_v.fill = FILL_ZINC_50
        cell_v.border = BORDER_ALL
    ws.row_dimensions[row].height = 44
    row += 1

    for i, (val, label) in enumerate(kpis):
        col = i + 1
        cell_l = ws.cell(row=row, column=col, value=label)
        cell_l.font = FONT_KPI_LABEL
        cell_l.alignment = ALIGN_CENTER
        cell_l.fill = FILL_ZINC_50
        cell_l.border = BORDER_ALL
    ws.row_dimensions[row].height = 22
    row += 2

    # Platform Breakdown
    ratings = data.get("ratings_summary", {})
    if ratings:
        row = _write_section_header(ws, row, "Platform Breakdown", 3)
        row = _write_header_row(ws, row, ["Platform", "Rating", "Reviews"])
        for platform, rating in ratings.items():
            _write_cell(ws, row, 1, platform, font=FONT_BODY_BOLD)
            _write_cell(ws, row, 2, rating, alignment=ALIGN_CENTER)
            _write_cell(ws, row, 3, "", alignment=ALIGN_CENTER)
            ws.row_dimensions[row].height = 22
            row += 1
        row += 1

    # Sentiment Distribution
    row = _write_section_header(ws, row, "Sentiment Distribution", 3)
    row = _write_header_row(ws, row, ["Sentiment", "Count", "Percentage"])
    for sentiment_key in ("positive", "mixed", "negative"):
        count = dist.get(sentiment_key, 0)
        pct = round(100 * count / total_sentiment) if total_sentiment else 0
        fill, font = _sentiment_style(sentiment_key)
        _write_cell(ws, row, 1, sentiment_key.title(), font=font, fill=fill, alignment=ALIGN_CENTER)
        _write_cell(ws, row, 2, count, alignment=ALIGN_CENTER, fill=fill, font=font)
        _write_cell(ws, row, 3, f"{pct}%", alignment=ALIGN_CENTER, fill=fill, font=font)
        row += 1
    row += 1

    # Overall Assessment
    summary = data.get("summary", "")
    if summary:
        row = _write_section_header(ws, row, "Overall Assessment", 6)
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=6)
        cell = ws.cell(row=row, column=1, value=summary)
        cell.font = FONT_BODY
        cell.alignment = ALIGN_WRAP
        cell.fill = FILL_ZINC_50
        cell.border = BORDER_ALL
        ws.row_dimensions[row].height = 60


def _build_theme_analysis(wb: Workbook, data: dict, company_name: str) -> None:
    ws = wb.create_sheet("Theme Analysis")
    _set_col_widths(ws, {1: 28, 2: 14, 3: 14, 4: 36, 5: 50})

    themes = data.get("themes", [])

    row = _write_title_bar(ws, 1, f"Theme Analysis — {company_name}", 5)
    row += 1

    if not themes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value="No themes identified.").font = FONT_BODY
        return

    row = _write_header_row(ws, row, ["Theme", "Frequency", "Sentiment", "Platforms", "Supporting Evidence"])

    reviews = data.get("reviews", []) + data.get("employee_reviews", [])

    for idx, t in enumerate(themes):
        theme_name = t.get("theme", "")
        freq = t.get("frequency", "low")
        sentiment = t.get("sentiment", "mixed")
        platforms = t.get("platforms", [])
        platforms_str = ", ".join(platforms) if isinstance(platforms, list) else str(platforms)

        # Find a supporting quote
        quote = ""
        theme_lower = theme_name.lower()
        for r in reviews:
            text = r.get("text", "")
            if any(word in text.lower() for word in theme_lower.split()[:2]):
                quote = f'"{text[:120]}..."' if len(text) > 120 else f'"{text}"'
                break

        bg = _alt_row_fill(idx)
        _write_cell(ws, row, 1, theme_name, font=FONT_BODY_BOLD, fill=bg)

        f_fill, f_font = _frequency_style(freq)
        _write_cell(ws, row, 2, freq.upper(), font=f_font, fill=f_fill, alignment=ALIGN_CENTER)

        s_fill, s_font = _sentiment_style(sentiment)
        _write_cell(ws, row, 3, sentiment.title(), font=s_font, fill=s_fill, alignment=ALIGN_CENTER)

        _write_cell(ws, row, 4, platforms_str, font=FONT_SMALL, fill=bg, alignment=ALIGN_LEFT_TOP)
        _write_cell(ws, row, 5, quote, font=FONT_SMALL, fill=bg, alignment=ALIGN_LEFT_TOP)

        ws.row_dimensions[row].height = 32
        row += 1


def _build_consumer_reviews(wb: Workbook, data: dict, company_name: str) -> None:
    ws = wb.create_sheet("Consumer Reviews")
    _set_col_widths(ws, {1: 16, 2: 10, 3: 65, 4: 14, 5: 14, 6: 30})

    reviews = data.get("reviews", [])

    row = _write_title_bar(ws, 1, f"Consumer Reviews — {company_name} ({len(reviews)} reviews)", 6)
    row += 1

    row = _write_header_row(ws, row, ["Platform", "Rating", "Review Text", "Date", "Sentiment", "Author"])

    # Freeze header
    ws.freeze_panes = f"A{row}"

    if not reviews:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="No consumer reviews found.").font = FONT_BODY
        return

    for idx, rev in enumerate(reviews):
        bg = _alt_row_fill(idx)
        _write_cell(ws, row, 1, rev.get("platform", ""), fill=bg)
        _write_cell(ws, row, 2, rev.get("rating", ""), alignment=ALIGN_CENTER, fill=bg)
        _write_cell(ws, row, 3, rev.get("text", ""), fill=bg)
        _write_cell(ws, row, 4, rev.get("date", ""), alignment=ALIGN_CENTER, fill=bg)

        sentiment = rev.get("sentiment", "mixed")
        s_fill, s_font = _sentiment_style(sentiment)
        _write_cell(ws, row, 5, sentiment.title(), font=s_font, fill=s_fill, alignment=ALIGN_CENTER)

        _write_cell(ws, row, 6, rev.get("author", ""), font=FONT_SMALL, fill=bg)
        ws.row_dimensions[row].height = 28
        row += 1

    # Auto-filter on header row
    ws.auto_filter.ref = f"A3:{get_column_letter(6)}{row - 1}"


def _build_employee_reviews(wb: Workbook, data: dict, company_name: str) -> None:
    ws = wb.create_sheet("Employee Reviews")
    _set_col_widths(ws, {1: 14, 2: 10, 3: 28, 4: 38, 5: 38, 6: 14, 7: 14})

    emp_reviews = data.get("employee_reviews", [])

    row = _write_title_bar(ws, 1, f"Employee Reviews — {company_name} ({len(emp_reviews)} reviews)", 7)
    row += 1

    row = _write_header_row(ws, row, ["Platform", "Rating", "Title", "Pros", "Cons", "Date", "Sentiment"])

    ws.freeze_panes = f"A{row}"

    if not emp_reviews:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws.cell(row=row, column=1, value="No employee reviews found.").font = FONT_BODY
        return

    for idx, er in enumerate(emp_reviews):
        bg = _alt_row_fill(idx)
        _write_cell(ws, row, 1, er.get("platform", ""), fill=bg)
        _write_cell(ws, row, 2, er.get("rating", ""), alignment=ALIGN_CENTER, fill=bg)
        _write_cell(ws, row, 3, er.get("title", ""), font=FONT_BODY_BOLD, fill=bg)
        _write_cell(ws, row, 4, er.get("pros", ""), font=FONT_SMALL, fill=bg)
        _write_cell(ws, row, 5, er.get("cons", ""), font=FONT_SMALL, fill=bg)
        _write_cell(ws, row, 6, er.get("date", ""), alignment=ALIGN_CENTER, fill=bg)

        sentiment = er.get("sentiment", "mixed")
        s_fill, s_font = _sentiment_style(sentiment)
        _write_cell(ws, row, 7, sentiment.title(), font=s_font, fill=s_fill, alignment=ALIGN_CENTER)

        ws.row_dimensions[row].height = 32
        row += 1

    ws.auto_filter.ref = f"A3:{get_column_letter(7)}{row - 1}"


def _build_recommendations(wb: Workbook, data: dict, company_name: str) -> None:
    ws = wb.create_sheet("Recommendations")
    _set_col_widths(ws, {1: 6, 2: 28, 3: 14, 4: 14, 5: 50, 6: 36})

    row = _write_title_bar(ws, 1, f"Recommendations — {company_name}", 6)
    row += 1

    themes = data.get("themes", [])
    action_themes = [t for t in themes if t.get("sentiment", "").lower() in ("negative", "mixed")]

    if not action_themes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="No action items identified — sentiment is predominantly positive.").font = FONT_BODY
        return

    row = _write_header_row(ws, row, ["#", "Issue", "Severity", "Frequency", "Recommendation", "Platforms"])

    for idx, t in enumerate(action_themes[:10]):
        theme_name = t.get("theme", "")
        freq = t.get("frequency", "low")
        sentiment = t.get("sentiment", "mixed")
        platforms = t.get("platforms", [])
        platforms_str = ", ".join(platforms) if isinstance(platforms, list) else str(platforms)

        severity = "High" if sentiment == "negative" and freq == "high" else (
            "Medium" if sentiment == "negative" or freq in ("high", "medium") else "Low"
        )

        recommendation = f"Address '{theme_name}' — reported as {sentiment} across {platforms_str} with {freq} frequency."

        bg = _alt_row_fill(idx)
        _write_cell(ws, row, 1, idx + 1, alignment=ALIGN_CENTER, font=FONT_BODY_BOLD, fill=bg)
        _write_cell(ws, row, 2, theme_name, font=FONT_BODY_BOLD, fill=bg)

        sev_map = {"High": (FILL_RED, FONT_RED), "Medium": (FILL_AMBER, FONT_AMBER), "Low": (FILL_GREEN, FONT_GREEN)}
        sv_fill, sv_font = sev_map.get(severity, (FILL_AMBER, FONT_AMBER))
        _write_cell(ws, row, 3, severity, font=sv_font, fill=sv_fill, alignment=ALIGN_CENTER)

        f_fill, f_font = _frequency_style(freq)
        _write_cell(ws, row, 4, freq.title(), font=f_font, fill=f_fill, alignment=ALIGN_CENTER)

        _write_cell(ws, row, 5, recommendation, font=FONT_BODY, fill=bg)
        _write_cell(ws, row, 6, platforms_str, font=FONT_SMALL, fill=bg)

        ws.row_dimensions[row].height = 36
        row += 1


# ======================================================================
# Public entry point
# ======================================================================


def generate_cx_xlsx(data: dict, company_name: str, output_path: Path | str) -> Path:
    """Build a 5-sheet CX Intelligence XLSX and save it to *output_path*."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    accent = DEFAULT_ACCENT
    wb = Workbook()

    _build_executive_summary(wb, data, company_name, accent)
    _build_theme_analysis(wb, data, company_name)
    _build_consumer_reviews(wb, data, company_name)
    _build_employee_reviews(wb, data, company_name)
    _build_recommendations(wb, data, company_name)

    wb.save(str(output_path))
    logger.info("CX XLSX saved to %s", output_path)
    return output_path
